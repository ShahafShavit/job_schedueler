from json_handler import *
import random
from os import remove
import openpyxl
def generate_dummy_employee_data(month, num_employees=10): # <<<<< DEPRECATED >>>>>>>
    filename = "employees.json"
    data = []

    # Fetch the number of venues from the venues file
    venues_data = load_data_from_json(f"{config.get_location('data')}venues_{month}.json")
    num_venues = len(venues_data)
    venue_names = [venue["name"] for venue in venues_data]

    for i in range(1, num_employees + 1):
        # Randomly assign 1 to 3 venues to each employee
        allowed_venues = random.sample(venue_names, random.randint(1, 3))

        data.append({
            "id": f"#E{i}",
            "allowed_venues": allowed_venues
        })

    save_data_to_json(filename, data)

def generate_dummy_shift_data(month):
    print("Generating dummy shift data")
    filename = f"{config.get_location('data')}shifts_{month}.json"
    data = []

    # Fetch the employee names from the employees file
    employees_data = load_data_from_json("employees.json")

    # Fetch the number of shifts per day from the settings
    shifts_per_day = config.get_default('default_shifts_per_day')

    for emp_data in employees_data:
        # Generate random number of unavailable dates for each employee
        num_unavailable_dates = random.randint(1, 8)
        unavailable_dates_data = {}

        for _ in range(num_unavailable_dates):
            date = f"{random.randint(1, 30)}/{month}"
            # Generate a list of boolean values for each shift
            shifts_unavailability = [random.choice([True, False]) for _ in range(shifts_per_day)]
            unavailable_dates_data[date] = shifts_unavailability

        data.append({
            "id": emp_data["id"],
            "unavailable_dates": unavailable_dates_data
        })

    save_data_to_json(filename, data)

def generate_dummy_venue_data(month):
    print("Generating dummy venue data")
    filename = f"{config.get_location('data')}venues_{month}.json"
    data = []

    # Fetch the venue names from the employees file
    employees_data = load_data_from_json("employees.json")
    venue_names = list(set(venue for emp in employees_data for venue in emp["allowed_venues"].keys()))

    # Get the number of shifts per day from settings
    shifts_per_day = config.get_default('default_shifts_per_day')

    for venue_name in venue_names:
        # Generate random number of event dates for each venue
        num_event_dates = random.randint(10, 14)
        event_dates_dict = {}

        for _ in range(num_event_dates):
            date = f"{random.randint(1, 30)}/{month}"
            # Assuming no morning events and all events are in the evening
            # Create a list of False values for all shifts except the last one (evening)
            shifts = [False] * (shifts_per_day - 1) + [True]
            event_dates_dict[date] = shifts

        # Sort the dates
        sorted_event_dates = dict(sorted(event_dates_dict.items(), key=lambda x: (int(x[0].split("/")[1]), int(x[0].split("/")[0]))))

        data.append({
            "name": venue_name,
            "event_dates": sorted_event_dates
        })

    save_data_to_json(filename, data)



def delete_files_for_month(month):
    files_to_delete = [
        f"{config.get_location('data')}venues_{month}.json",
        f"{config.get_location('data')}shifts_{month}.json",
        f"{config.get_location('output')}schedule_{month}.json"
    ]

    for filename in files_to_delete:
        try:
            remove(filename)
            print(f"Deleted {filename}")
        except FileNotFoundError:
            print(f"{filename} not found.")
        except Exception as e:
            print(f"Error deleting {filename}: {e}")


def generate_dummy_unavailability_in_excel(month, year):
    print("Generating dummy unavailability in excel")
    # Load the Excel file
    filename = f"{config.get_location('calendar_export')}Hagashot_{month}_{year}.xlsx"
    wb = openpyxl.load_workbook(filename)

    # Fetch the number of shifts per day from the settings
    shifts_per_day = config.get_default('default_shifts_per_day')

    # Using the first worksheet to determine the potential cells
    ws = wb.worksheets[0]

    # Create a list of all potential cells that can be marked as unavailable
    potential_cells = []

    # Start from column E (5) and row 6
    for col_num in range(5, 12):  # Columns for days of the week
        for row_num in range(6, ws.max_row + 1):
            cell = ws.cell(row=row_num, column=col_num)
            # Check if the cell contains an integer (date)
            if isinstance(cell.value, int):
                # Check the cells below it to see if they form a valid calendar cell
                for shift in range(1, shifts_per_day + 1):
                    below_cell = ws.cell(row=row_num + shift, column=col_num)
                    # Ensure the cell is empty
                    if not below_cell.value:
                        potential_cells.append(below_cell)

    # For each sheet (employee) in the workbook
    for ws in wb.worksheets:
        # Fill each potential cell with 'No' for testing purposes
        for cell in potential_cells:
            if random.random() < 0.2:
                ws.cell(row=cell.row, column=cell.column).value = 'No'

    # Save the modified Excel file
    wb.save(filename)







