from openpyxl.utils import get_column_letter

from json_handler import *
from calendar import monthrange, weekday, SUNDAY
from openpyxl.styles import Font,Alignment,Border,Side
from openpyxl import Workbook, load_workbook

def column_auto_width(ws):
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    # Create an alignment object with horizontal alignment set to "center"
    center_aligned = Alignment(horizontal="center")

    # Loop through all the cells in the worksheet and apply the alignment
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center_aligned

def set_column_widths(worksheet, start_col, end_col, width):
    """
    Set the width of a range of columns in an Excel worksheet.
    """

    for col_idx in range(start_col, end_col + 1):
        col_letter = get_column_letter(col_idx)
        worksheet.column_dimensions[col_letter].width = width

def add_borders_to_calendar_cell(worksheet, start_row_idx, col_idx, shifts_per_day):
    thin_border = Border(top=Side(style='thin'),
                         right=Side(style='thin'),
                         bottom=Side(style='thin'),
                         left=Side(style='thin'))

    # Top cell gets top and side borders
    top_cell = worksheet.cell(row=start_row_idx, column=col_idx)
    top_cell.border = Border(top=thin_border.top,
                             right=thin_border.right,
                             left=thin_border.left)

    # Middle cells get only side borders
    for shift in range(1, shifts_per_day):
        middle_cell = worksheet.cell(row=start_row_idx + shift, column=col_idx)
        middle_cell.border = Border(right=thin_border.right,
                                    left=thin_border.left)

    # Bottom cell gets bottom and side borders
    bottom_cell = worksheet.cell(row=start_row_idx + shifts_per_day, column=col_idx)
    bottom_cell.border = Border(bottom=thin_border.bottom,
                                right=thin_border.right,
                                left=thin_border.left)



def report_work_by_date_to_excel(month, wb):
    schedule_data = load_data_from_json(f"{config.get_location('output')}schedule_{month}.json")
    ws = wb.create_sheet(title="Work By Date")

    # Create a dictionary to group assignments by date
    grouped_by_date = {}
    # Create a dictionary to count shifts for each employee
    shifts_count = {}

    for venue, dates in schedule_data.items():
        if venue != "shifts":
            for date, employee_id in dates.items():
                if date not in grouped_by_date:
                    grouped_by_date[date] = []
                grouped_by_date[date].append((venue, employee_id))

                # Count shifts for each employee
                shifts_count[employee_id] = shifts_count.get(employee_id, 0) + 1

    # Add headers
    ws.append(["Date", "Venue", "Employee ID"])

    # Write data to the worksheet
    for date, assignments in sorted(grouped_by_date.items(), key=lambda x: list(map(int, x[0].split("/")))):
        row_data = [date]
        for venue, employee_id in assignments:
            row_data.extend([venue, employee_id])
        ws.append(row_data)

    ws.append([])  # Add an empty row for separation

    # Write the number of shifts each employee received
    ws.append(["Shifts count:"])
    for employee_id, count in shifts_count.items():
        ws.append([employee_id, f"{count} shifts"])
    column_auto_width(ws)

def report_venue_dates_with_workers_to_excel(month, wb):
    schedule_data = load_data_from_json(f"{config.get_location('output')}schedule_{month}.json")
    ws = wb.create_sheet(title="Venue Dates With Workers")

    for venue, dates in schedule_data.items():
        ws.append([f"Venue: {venue}"])
        for date, worker in dates.items():
            ws.append([date, worker])
        ws.append(["------"])
    column_auto_width(ws)


def report_worker_shifts_to_excel(month, wb):
    schedule_data = load_data_from_json(f"{config.get_location('output')}schedule_{month}.json")

    worker_shifts = {}
    for venue, dates in schedule_data.items():
        for date, worker in dates.items():
            if worker not in worker_shifts:
                worker_shifts[worker] = []
            worker_shifts[worker].append((date, venue))

    # Add a new worksheet to the provided workbook
    ws = wb.create_sheet(title="Worker Shifts")

    # Add headers
    ws['A1'] = 'Worker'
    ws['B1'] = 'Date'
    ws['C1'] = 'Venue'
    ws['A1'].font = ws['B1'].font = ws['C1'].font = Font(bold=True)

    row = 2
    for worker, shifts in worker_shifts.items():
        for date, venue in sorted(shifts, key=lambda x: list(map(int, x[0].split("/")))):
            ws[f'A{row}'] = worker
            ws[f'B{row}'] = date
            ws[f'C{row}'] = venue
            row += 1
    column_auto_width(ws)


def report_employee_availability_to_excel(month, wb):
    # Load employee data, shifts data, and venues data
    employees = load_data_from_json("employees.json")
    shifts_data = load_data_from_json(f"{config.get_location('data')}shifts_{month}.json")
    venues = load_data_from_json(f"{config.get_location('data')}venues_{month}.json")

    # Create a dictionary to map employee IDs to their unavailable dates
    unavailable_dates_map = {shift["id"]: shift["unavailable_dates"] for shift in shifts_data}
    # Determine the last day of the month
    _, last_day = monthrange(2023, int(month))  # Assuming the year is 2023

    # Generate the date range
    dates = [f"{day}/{month}" for day in range(1, last_day + 1)]

    # Create a new sheet in the workbook
    ws = wb.create_sheet(title="Employee Availability")

    # Set dates as column headers
    ws.append(["Employee"] + dates)

    # Fill in employee availability
    for employee in employees:
        row_data = [employee["id"]]
        for date in dates:
            # Check if employee is available on this date using the unavailable_dates_map
            if date in unavailable_dates_map.get(employee["id"], []):
                row_data.append("")  # Employee is unavailable
            else:
                row_data.append(employee["id"])  # Employee is available
        ws.append(row_data)

    # Add a blank row after employee availability data
    ws.append([])

    # Add venue event dates
    for venue in venues:
        venue_name = venue["name"]
        row_data = [venue_name]
        for date in dates:
            if date in venue["event_dates"]:
                row_data.append(venue_name)
            else:
                row_data.append("")
        ws.append(row_data)

    column_auto_width(ws)

def generate_excel_reports(month):
    wb = Workbook()
    wb.remove(wb.active)  # Remove the default sheet

    report_worker_shifts_to_excel(month, wb)
    report_work_by_date_to_excel(month, wb)
    report_venue_dates_with_workers_to_excel(month, wb)
    report_employee_availability_to_excel(month, wb)

    # Ensure the directory exists
    makedirs(config.get_location('excel_reports'), exist_ok=True)

    # Save the workbook to a file
    wb.save(f"{config.get_location('excel_reports')}reports_{month}.xlsx")

def generate_monthly_calendar_excel(month, year):
    # Load employee data and settings
    employees = load_data_from_json("employees.json")
    settings = load_data_from_json("settings.json")
    shifts_per_day = config.get_default('default_shifts_per_day')
    shift_names = config.get_default("default_shift_names")

    # Create a new workbook
    wb = Workbook()

    # Remove the default sheet created
    wb.remove(wb.active)

    # Hebrew names for the days of the week in the correct order
    hebrew_days = ["ראשון", "שני", "שלישי", "רביעי", "חמישי", "שישי", "שבת"]

    # Loop through each employee
    for employee in employees:
        employee_name = employee["id"]

        # Create a new sheet for the employee
        ws = wb.create_sheet(title=employee_name)

        # Set the sheet to RTL
        ws.sheet_view.rightToLeft = True

        # Merge cells and set values
        ws.merge_cells('E1:K2')
        ws['E1'].value = f'הגשות עבור {employee_name}'
        ws.merge_cells('E3:K4')
        ws['E3'].value = f'{month}/{year}'

        # Set the headers for the days of the week starting with Sunday
        for idx, day in enumerate(hebrew_days, start=5):  # Starting from the 5th column
            ws.cell(row=5, column=idx).value = day

        # Determine the first day of the month and the last day of the month
        first_weekday, last_day = monthrange(year, int(month))

        # Adjust the weekday to start with Sunday
        first_weekday = (first_weekday + 1) % 7

        # Initialize variables to keep track of the current date and current weekday
        current_date = 1
        current_weekday = first_weekday

        # Loop through the weeks of the month
        row_start = 6  # Starting from the 6th row
        while current_date <= last_day:
            # Create a row for the dates
            date_row = ["" for _ in range(7)]
            # Fill in the dates for this week
            for i in range(current_weekday, 7):
                if current_date <= last_day:
                    date_row[i] = current_date
                    current_date += 1
            for idx, date in enumerate(date_row, start=5):  # Starting from the 5th column
                ws.cell(row=row_start, column=idx).value = date
                add_borders_to_calendar_cell(ws, row_start, idx, shifts_per_day)

            # Add shift names to the side of the calendar
            for idx, shift in enumerate(shift_names):
                ws.cell(row=row_start + idx + 1, column=4).value = shift  # +1 to start from the 7th row

            # Move to the next set of rows for the next week
            row_start += shifts_per_day + 1

            # Reset the current weekday for the next week
            current_weekday = 0

        set_column_widths(ws, 5, 11, 15)

        center_alignment = Alignment(horizontal='center', vertical='center')

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = center_alignment

    makedirs(config.get_location('calendar_export'), exist_ok=True)
    # Save the workbook to a file
    wb.save(f"{config.get_location('calendar_export')}Hagashot_{month}_{year}.xlsx")


def extract_unavailability_from_excel(month, year):
    print("Extracting unavailability from excel")
    # Construct the path to the workbook
    excel_file_path = f"{config.get_location('calendar_export')}Hagashot_{month}_{year}.xlsx"
    wb = load_workbook(excel_file_path)

    # Load settings
    shifts_per_day = config.get_default('default_shifts_per_day')

    # List to store unavailability data for all employees
    all_unavailability_data = []

    # Iterate over each sheet (employee)
    for ws in wb.worksheets:
        # Extract employee name from the sheet title
        employee_name = ws.title

        # Determine the first day of the month and the last day of the month
        _, last_day = monthrange(int(year), int(month))

        # Dictionary to store unavailability data for this employee
        employee_unavailability = {
            "id": employee_name,
            "unavailable_dates": {}
        }

        # Start reading from the 6th row
        row = 6

        while row < ws.max_row:
            # Read the dates for the week
            week_dates = [ws.cell(row=row, column=col).value for col in range(5, 12)]

            # For each shift, check the unavailability
            for shift in range(1, shifts_per_day + 1):
                row += 1
                for col, date in enumerate(week_dates, start=5):
                    if ws.cell(row=row, column=col).value == "No":
                        if f"{date}/{month}" not in employee_unavailability["unavailable_dates"]:
                            employee_unavailability["unavailable_dates"][f"{date}/{month}"] = [False] * shifts_per_day
                        employee_unavailability["unavailable_dates"][f"{date}/{month}"][shift - 1] = True

            # Move to the next week's dates
            row += 1

        all_unavailability_data.append(employee_unavailability)

    save_data_to_json(f"{config.get_location('data')}shifts_{month}.json",all_unavailability_data)


